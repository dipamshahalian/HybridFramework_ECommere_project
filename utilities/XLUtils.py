import openpyxl
from openpyxl.styles import PatternFill

def getRowCount(file,sheetName):
    wb = openpyxl.load_workbook(file)
    sheet = wb[sheetName]
    return (sheet.max_row)

def getColumnCount(file,sheetName):
    wb = openpyxl.load_workbook(file)
    sheet = wb[sheetName]
    return(sheet.max_column)

def readData(file,sheetName,rownum,columnnum):
    wb = openpyxl.load_workbook(file)
    sheet = wb[sheetName]
    return sheet.cell(rownum,columnnum).value

def writeData(file,sheetName,rownum,columnnum,data):
    wb = openpyxl.load_workbook(file)
    sheet = wb[sheetName]
    sheet.cell(rownum,columnnum).value= data
    wb.save(file)

def fillGreenColor(file,sheetName,rownum,columnnum):
    wb = openpyxl.load_workbook(file)
    sheet = wb[sheetName]
    greenFill = PatternFill(start_color='60b212',
                            end_color='60b212',
                            fill_type='solid')
    sheet.cell(rownum,columnnum).fill=greenFill
    wb.save(file)

def fillRedColor(file,sheetName,rownum,columnnum):
    wb = openpyxl.load_workbook(file)
    sheet = wb[sheetName]
    redFill = PatternFill(start_color='ff0000',
                            end_color='ff0000',
                            fill_type='solid')
    sheet.cell(rownum, columnnum).fill = redFill
    wb.save(file)

